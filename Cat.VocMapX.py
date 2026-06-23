import pandas as pd
import csv
import xml.etree.ElementTree as ET
import html
import re
import os
import pickle
import joblib
import multiprocessing
import requests
import time
import numpy as np
import torch
import faiss
import psutil
import gc
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from tqdm import tqdm
from rapidfuzz import process, fuzz
from sklearn.feature_extraction.text import TfidfVectorizer
from sentence_transformers import SentenceTransformer

# =============================================================================
# CONFIGURATION
# =============================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
VOCAB_DIR = os.path.join(BASE_DIR, "Vocabulaires")

# Caches
INDEX_CACHE      = os.path.join(VOCAB_DIR, "thesaurus_index.pkl")
PUBCHEM_CACHE    = os.path.join(VOCAB_DIR, "pubchem_api_cache.pkl")
TFIDF_CACHE      = os.path.join(VOCAB_DIR, "tfidf_vectorizer.pkl")
FAISS_INDEX_PATH = os.path.join(VOCAB_DIR, "faiss_index_bge.bin")

# Fichiers I/O
FILE_KEYWORDS    = 'liste_mots_cles_uniques.csv'
FILE_OUTPUT_CSV  = 'alignement_3etapes_complet.csv'
FILE_STATS_XLSX  = 'statistiques_vocabulaires_seuils.xlsx'

# --- Paramètres Étape 2 : Fuzzy (Voc_MapX.py) ---
THRESHOLDS  = [100, 98, 97, 96, 95]
SCORE_CUTOFF = min(THRESHOLDS) - 1   # 89
TFIDF_TOP_K  = 200
MAX_WORKERS  = min(4, multiprocessing.cpu_count())
FUZZY_THRESHOLD_FOR_SEMANTIC = 98

# --- Paramètres Étape 3 : Sémantique (Voc_MapX_FAISS_LLM.py) ---
# --- Modèle prenant en compte le multilangue ---
MODEL_NAME = 'BAAI/bge-m3'
BGE_QUERY_PREFIX = "Represent this science keyword for retrieving relevant thesaurus terms: "
# Seuils L2
DIST_THRESHOLDS = [0.3, 0.5, 0.7, 0.8]

# --- Paramètres Étape 4 : Vérification LLM (Ollama/Mistral) ---
OLLAMA_URL       = "http://localhost:11434/api/generate"
OLLAMA_MODEL     = "mistral:7b"
OLLAMA_TIMEOUT   = 60          # secondes par requête
OLLAMA_WORKERS   = 2           # parallélisme modéré pour ne pas saturer la RAM
FILE_LLM_REPORT  = 'rapport_verification_llm.xlsx'
LLM_CACHE        = os.path.join(VOCAB_DIR, "llm_verification_cache.pkl")

# --- Paramètres Étape 5 : LLM sur NON TROUVÉ ---
TOP_N_CANDIDATES  = 10         # nombre de candidats FAISS soumis à Mistral
FILE_LLM_STEP5    = 'rapport_llm_non_trouves.xlsx'
LLM_CACHE_STEP5   = os.path.join(VOCAB_DIR, "llm_step5_cache.pkl")

# Vocabulaires sources
FILES = {
    'PUBCHEM':       (None,                                    'api'),
    'BLH':           ('BLH.rdf',                               'rdf'),
    'AGROVOC':       ('agrovoc_lod.nt',                        'nt'),
    'EUROVOC':       ('eurovoc-skos-ap-eu.rdf',                'rdf'),
    'GEMET':         ('gemet.rdf',                             'rdf'),
    'MESH':          ('2024_mesh_20250722.nq',                  'nq'),
    'CHEBI':         ('chebi.sdf',                             'sdf'),
    'GOLDBOOK':      ('goldbook_terms.xml',                    'xml'),
    'NALT':          ('nalt-full_dwn_20240716.rdf',            'rdf'),
    'UNESCO_SOIL':   ('UNESCO-Thesaurus-Soil-Science.rdf',     'rdf'),
    'EDAM_MAIN':     ('EDAM_1.25.owl',                         'rdf'),
    'EDAM_IMAGING':  ('EDAM-bioimaging_alpha06.owl',           'rdf'),
    'PACTOLS_LIEUX': ('pactols_lieux_all_20260123.rdf',        'rdf'),
    'PACTOLS_SUJETS':('pactols_sujets_all_20260123.rdf',       'rdf'),
    'ENVO':          ('envo.owl',                              'owl'),
    'PO':            ('po.owl',                                'owl'),
}


# =============================================================================
# UTILITAIRE : conversion sécurisée des embeddings vers numpy CPU
# (sentence-transformers peut retourner un tenseur torch OU une liste selon
#  la version et le device — cette fonction gère les deux cas)
# =============================================================================

def to_numpy_cpu(emb):
    """Convertit un résultat de model.encode() en np.ndarray float32 CPU."""
    import torch
    if isinstance(emb, torch.Tensor):
        return emb.detach().cpu().numpy().astype('float32')
    elif isinstance(emb, list):
        return np.array(emb, dtype='float32')
    else:
        # déjà un np.ndarray
        return emb.astype('float32')

# =============================================================================
# UTILITAIRES SYSTÈME (Étape 3)
# =============================================================================

def get_optimal_batch_size():
    """Ajuste le batch_size selon la RAM disponible."""
    available_ram = psutil.virtual_memory().available / (1024 ** 3)
    if available_ram > 16: return 10
    elif available_ram > 8: return 8
    else: return 4

def print_ram_usage(step_name):
    process = psutil.Process(os.getpid())
    mem = process.memory_info().rss / (1024 ** 3)
    print(f"  [RAM] {step_name} : {mem:.2f} Go utilisés")

# =============================================================================
# FONCTIONS API PUBCHEM
# =============================================================================

def query_pubchem(keyword, timeout=10):
    """Interroge l'API PubChem pour une correspondance exacte."""
    url = f"https://pubchem.ncbi.nlm.nih.gov/rest/rdf/query?graph=synonym&name={keyword}&format=json"
    try:
        time.sleep(0.3)
        response = requests.get(url, timeout=timeout)
        if response.status_code == 200:
            data = response.json()
            bindings = data.get("results", {}).get("bindings", [])
            for b in bindings:
                val = b.get("value", {}).get("value", "")
                if val.lower() == str(keyword).lower():
                    return b.get("synonym", {}).get("value")
    except Exception:
        pass
    return None

# =============================================================================
# PARSEURS LOCAUX (Voc_MapX.py — inchangés)
# =============================================================================

def parse_xml_robust(file_path, source_name):
    lookup = {}
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        target_tags = {'prefLabel', 'altLabel', 'label', 'title', 'hasExactSynonym', 'hasNarrowSynonym'}
        for elem in root.iter():
            uri = next((v for k, v in elem.items() if k.endswith(('about', 'resource'))), "")
            for child in elem:
                tag_local = child.tag.split('}')[-1]
                if tag_local in target_tags:
                    lang = child.get('{http://www.w3.org/XML/1998/namespace}lang')
                    if lang in [None, 'fr', 'en']:
                        txt = child.text.strip() if child.text else None
                        if txt:
                            lookup[txt.lower()] = (txt, uri, source_name, tag_local)
    except Exception as e:
        print(f"  Erreur {source_name}: {e}")
    return lookup

def parse_mesh_nq(file_path):
    lookup = {}
    regex = re.compile(r'<(.*?)>\s+<.*#(prefLabel|altLabel|label)>\s+"(.*?)"@([a-z]{2})')
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                m = regex.search(line)
                if m and m.group(4) in ['fr', 'en']:
                    lookup[m.group(3).lower()] = (m.group(3), m.group(1), "MESH", m.group(2))
    except Exception:
        pass
    return lookup

def parse_obo_owl(file_path, source_name):
    """
    Parse les ontologies OBO au format OWL/RDF (ENVO, PO, GO, etc.).
    Filtre les concepts par préfixe URI pour exclure les métadonnées OWL,
    et nettoie les synonymes qui embarquent la langue entre parenthèses
    (ex: "pollen sac wall (exact)", "planta entera (Spanish, exact)").
    """
    lookup = {}
    # Préfixe URI attendu : "ENVO_", "PO_", etc. — dérivé du source_name
    uri_prefix = source_name + '_'
    target_tags = {'label', 'prefLabel', 'altLabel',
                   'hasExactSynonym', 'hasNarrowSynonym',
                   'hasRelatedSynonym', 'hasBroadSynonym'}
    # Détecte les suffixes de langue OBO : "(Spanish, exact)", "(Japanese, broad)", "(exact)", etc.
    lang_suffix = re.compile(
        r'\s*\((spanish|japanese|french|german|exact|broad|narrow|related)[^)]*\)\s*$',
        re.IGNORECASE
    )
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        for elem in root.iter():
            uri = next(
                (v for k, v in elem.items() if k.endswith(('about', 'resource'))),
                ''
            )
            # Garder uniquement les vrais concepts de l'ontologie
            if uri_prefix not in uri:
                continue
            for child in elem:
                tag_local = child.tag.split('}')[-1]
                if tag_local not in target_tags:
                    continue
                lang = child.get('{http://www.w3.org/XML/1998/namespace}lang')
                # Accepter : pas d'attribut lang (OBO), 'en', 'fr'
                # Rejeter : 'es', 'ja', 'de', etc.
                if lang not in (None, 'en', 'fr'):
                    continue
                txt = child.text.strip() if child.text else None
                if not txt:
                    continue
                # Décoder les entités HTML (ex: &#243; → ó dans PO)
                txt = html.unescape(txt)
                # Supprimer les suffixes de langue OBO embarqués dans le texte
                txt = lang_suffix.sub('', txt).strip()
                # Rejeter les entrées non-ASCII (japonais, etc.)
                # et les entités HTML malformées non résolues (ex: &#n)
                if not txt.isascii() or '&#' in txt:
                    continue
                if txt:
                    key = txt.lower()
                    if key not in lookup or tag_local in ('label', 'prefLabel'):
                        lookup[key] = (txt, uri, source_name, tag_local)
    except Exception as e:
        print(f"  Erreur {source_name} : {e}")
    return lookup

def parse_agrovoc_nt(file_path):
    """
    Parse le fichier AGROVOC au format N-Triples (.nt).
    Extrait les prefLabel et altLabel en français et en anglais.
    Le format NT est : <sujet> <prédicat> <objet> .
    L'objet littéral est de la forme "texte"@lang ou "texte"^^type.
    """
    lookup = {}
    # Capture : URI sujet | type de label | texte | langue
    regex = re.compile(
        r'<([^>]+)>\s+<[^>]*#(prefLabel|altLabel|hiddenLabel)>\s+"((?:[^"\\]|\\.)*)"\s*@([a-z]{2,3})\s*\.'
    )
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                m = regex.search(line)
                if not m:
                    continue
                uri, tag, text, lang = m.group(1), m.group(2), m.group(3), m.group(4)
                if lang not in ('fr', 'en'):
                    continue
                text = text.strip()
                if not text:
                    continue
                # Les prefLabel ont priorité sur altLabel : on n'écrase pas un prefLabel déjà indexé
                key = text.lower()
                if key not in lookup or tag == 'prefLabel':
                    lookup[key] = (text, uri, 'AGROVOC', tag)
    except Exception as e:
        print(f"  Erreur AGROVOC : {e}")
    return lookup

def parse_chebi_sdf(file_path):
    lookup = {}
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            records = f.read().split('$$$$')
            for r in records:
                id_m = re.search(r'> <ChEBI ID>\n(CHEBI:\d+)', r)
                if id_m:
                    chebi_id = id_m.group(1)
                    for tag in ['ChEBI NAME', 'IUPAC_NAME', 'SYNONYM']:
                        names = re.findall(rf'> <{tag}>\n(.*?)\n', r)
                        for n in names:
                            if n.strip():
                                lookup[n.strip().lower()] = (n.strip(), chebi_id, "CHEBI", tag)
    except Exception:
        pass
    return lookup

# =============================================================================
# TF-IDF (Étape 2 — Voc_MapX.py — inchangé)
# =============================================================================

def build_tfidf_vectorizer(refs):
    """Construit et retourne uniquement le vectoriseur TF-IDF ajusté sur les refs."""
    print("  Construction de l'index TF-IDF (n-grammes de caractères)...")
    # min_df=3 : supprime les n-grammes trop rares (frequence < 3 dans le corpus)
    # Reduit le vocabulaire de 40-60% sur grands corpus (500k+ termes) sans
    # degrader la qualite des candidats fuzzy, et divise par ~3 la taille du cache.
    vectorizer = TfidfVectorizer(analyzer='char_wb', ngram_range=(2, 3), min_df=3)
    vectorizer.fit(refs)
    print(f"  Vocabulaire TF-IDF : {len(vectorizer.vocabulary_)} n-grammes.")
    return vectorizer

def get_tfidf_candidates(kw_l, vectorizer, ref_matrix, refs, top_k=TFIDF_TOP_K):
    """
    Retourne les top_k candidats les plus proches de kw_l selon la similarité cosinus.
    Utilise un produit matriciel sparse pour éviter la densification complète en RAM.
    """
    kw_vec = vectorizer.transform([kw_l])
    sims = (ref_matrix @ kw_vec.T).toarray().flatten()
    top_idx = np.argpartition(sims, -min(top_k, len(refs)))[-min(top_k, len(refs)):]
    return [refs[i] for i in top_idx]

# =============================================================================
# VARIABLES GLOBALES DES WORKERS (Étape 2)
# =============================================================================

_worker_index      = None
_worker_refs       = None
_worker_vectorizer = None
_worker_ref_matrix = None

def _init_worker(index, refs, vectorizer):
    """
    Initialise les variables globales dans chaque worker.
    La matrice sparse est reconstruite localement pour éviter sa sérialisation.
    """
    global _worker_index, _worker_refs, _worker_vectorizer, _worker_ref_matrix
    _worker_index      = index
    _worker_refs       = refs
    _worker_vectorizer = vectorizer
    _worker_ref_matrix = vectorizer.transform(refs)

# =============================================================================
# WORKER FUZZY (Étape 2 — Voc_MapX.py — inchangé)
# =============================================================================

def match_worker_fuzzy(chunk, thresholds, pubchem_map, index=None, refs=None, vectorizer=None, ref_matrix=None):
    """
    Aligne un chunk de mots-clés contre l'index maître via TF-IDF + WRatio.
    Ne traite que les mots-clés non trouvés à l'étape exact match.
    Retourne un dict : kw -> résultats par seuil.
    Compatible spawn (macOS) : reçoit index/refs/vectorizer/ref_matrix en argument direct.
    ref_matrix est précalculée une seule fois avant la boucle (macOS) pour éviter
    de reconstruire la matrice sparse à chaque chunk (~500k termes × n-grammes).
    """
    if index is None:
        index      = _worker_index
        refs       = _worker_refs
        vectorizer = _worker_vectorizer
        ref_matrix = _worker_ref_matrix
    elif ref_matrix is None:
        # Fallback : construction à la volée si non fournie (ne devrait pas arriver)
        ref_matrix = vectorizer.transform(refs)

    results = {}
    for kw in chunk:
        kw_s = str(kw).strip()
        kw_l = kw_s.lower()
        row = {}

        # Pré-filtrage TF-IDF puis WRatio
        best_fuz = None
        if refs and kw_l not in index:
            candidates = get_tfidf_candidates(kw_l, vectorizer, ref_matrix, refs)
            best_fuz = process.extractOne(kw_l, candidates, scorer=fuzz.WRatio, score_cutoff=SCORE_CUTOFF)

        for t in thresholds:
            lbl, uid, src, sc = "NON TROUVÉ", "", "NON TROUVÉ", 0.0
            if kw_l in index:
                lbl, uid, src, _ = index[kw_l]; sc = 100.0
            elif best_fuz and best_fuz[1] >= t:
                lbl, uid, src, _ = index[best_fuz[0]]; sc = round(best_fuz[1], 1)
            row[f'Label_{t}']  = lbl
            row[f'ID_{t}']     = uid
            row[f'Source_{t}'] = src
            row[f'Score_{t}']  = sc

        results[kw_s] = row
    return results

# =============================================================================
# NORMALISATION POUR RECHERCHE SÉMANTIQUE (Étape 3 — Voc_MapX_FAISS_LLM.py)
# =============================================================================

def clean_term(text):
    if not text: return ""
    text = text.lower().replace('-', ' ')
    text = re.sub(r'[^a-z0-9\s]', '', text)
    text = " ".join(text.split())
    if text.endswith('s') and len(text) > 4:
        text = text[:-1]
    return text

def clean_keyword(kw):
    """
    Supprime les préfixes numériques de type '01. ', '09.02 ', '03.02 ', etc.
    Exemples :
        '01. ENVIRONNEMENT'              → 'ENVIRONNEMENT'
        '09.02 FLUID AND SOLID MECHANICS'→ 'FLUID AND SOLID MECHANICS'
        '07. SOCIAL SCIENCES'            → 'SOCIAL SCIENCES'
    """
    return re.sub(r'^\d{2}(?:\.\d{2})?\s*\.\s*', '', kw).strip()
# =============================================================================
# ÉTAPE 4 : VÉRIFICATION LLM VIA OLLAMA
# =============================================================================

def check_ollama_available():
    """Vérifie qu'Ollama tourne et que le modèle est disponible."""
    try:
        r = requests.get("http://localhost:11434/api/tags", timeout=5)
        if r.status_code == 200:
            models = [m['name'] for m in r.json().get('models', [])]
            available = any(OLLAMA_MODEL.split(':')[0] in m for m in models)
            if not available:
                print(f"  ⚠️  Modèle '{OLLAMA_MODEL}' non trouvé. Modèles disponibles : {models}")
                print(f"       Lancez : ollama pull {OLLAMA_MODEL}")
            return available
        return False
    except Exception:
        print("  ⚠️  Ollama n'est pas accessible sur localhost:11434.")
        print("       Assurez-vous qu'Ollama est lancé : ollama serve")
        return False

def verify_match_ollama(keyword, label, source, etape):
    """
    Interroge Mistral:7b via Ollama pour valider une paire (mot-clé, label).
    Retourne un dict avec le verdict, la confiance et l'explication.
    """
    prompt = f"""Tu es un expert en terminologie. Évalue si la correspondance suivante est sémantiquement valide.

Mot-clé : "{keyword}"
Label correspondant : "{label}"
Vocabulaire source : {source}
Méthode d'alignement : {etape}

Réponds UNIQUEMENT avec ces 3 lignes, sans aucun texte supplémentaire :
VERDICT: YES ou NO
CONFIDENCE: HIGH ou MEDIUM ou LOW
REASON: une phrase expliquant ta décision"""

    try:
        response = requests.post(
            OLLAMA_URL,
            json={"model": OLLAMA_MODEL, "prompt": prompt, "stream": False},
            timeout=OLLAMA_TIMEOUT
        )
        if response.status_code != 200:
            return {'verdict': 'ERROR', 'confidence': 'N/A', 'reason': f'HTTP {response.status_code}'}

        text = response.json().get("response", "").strip()

        # Parsing robuste ligne par ligne
        verdict    = "UNKNOWN"
        confidence = "UNKNOWN"
        reason     = ""
        for line in text.splitlines():
            line = line.strip()
            if line.upper().startswith("VERDICT:"):
                val = line.split(":", 1)[1].strip().upper()
                verdict = "YES" if "YES" in val else ("NO" if "NO" in val else "UNKNOWN")
            elif line.upper().startswith("CONFIDENCE:"):
                val = line.split(":", 1)[1].strip().upper()
                if val == "HIGH":        confidence = "HIGH"
                elif val == "MEDIUM":    confidence = "MEDIUM"
                elif val == "LOW":       confidence = "LOW"
                elif "MEDIUM" in val:    confidence = "MEDIUM"
                elif "HIGH" in val:      confidence = "HIGH"
                elif "LOW" in val:       confidence = "LOW"
            elif line.upper().startswith("REASON:"):
                reason = line.split(":", 1)[1].strip()

        return {'verdict': verdict, 'confidence': confidence, 'reason': reason}

    except requests.exceptions.Timeout:
        return {'verdict': 'ERROR', 'confidence': 'N/A', 'reason': 'Timeout Ollama'}
    except Exception as e:
        return {'verdict': 'ERROR', 'confidence': 'N/A', 'reason': str(e)}


def run_llm_verification(df_results, llm_cache):
    """
    Lance la vérification LLM sur toutes les paires trouvées (hors NON TROUVÉ).
    Utilise un cache pour éviter de revérifier les paires déjà traitées.
    Retourne le DataFrame enrichi + les données du rapport détaillé.
    """
    # Construire la liste des paires à vérifier
    pairs_to_verify = []
    for _, row in df_results.iterrows():
        kw    = row['Mot-clé original']
        etape = row['Etape_alignement']
        if etape == 'NON TROUVÉ':
            continue

        # Récupérer le meilleur label selon l'étape
        if etape == 'Exact':
            label  = row.get(f'Label_{THRESHOLDS[0]}', '')
            source = row.get(f'Source_{THRESHOLDS[0]}', '')
        elif etape == 'Fuzzy':
            # Prendre le seuil le plus haut qui a trouvé quelque chose
            label, source = '', ''
            for t in THRESHOLDS:
                if row.get(f'Source_{t}', 'NON TROUVÉ') != 'NON TROUVÉ':
                    label  = row.get(f'Label_{t}', '')
                    source = row.get(f'Source_{t}', '')
                    break
        elif etape == 'Sémantique':
            # Prendre le seuil sémantique le plus bas qui a trouvé quelque chose
            label, source = '', ''
            for thr in DIST_THRESHOLDS:
                col = str(thr).replace('.', '')
                if row.get(f'Source_sem_{col}', 'NON TROUVÉ') != 'NON TROUVÉ':
                    label  = row.get(f'Label_sem_{col}', '')
                    source = row.get(f'Source_sem_{col}', '')
                    break

        if label and label != 'NON TROUVÉ':
            cache_key = f"{kw}||{label}||{source}"
            pairs_to_verify.append({
                'kw': kw, 'label': label, 'source': source,
                'etape': etape, 'cache_key': cache_key
            })

    # Filtrer les paires non encore dans le cache
    new_pairs = [p for p in pairs_to_verify if p['cache_key'] not in llm_cache]
    already_cached = len(pairs_to_verify) - len(new_pairs)

    print(f"\n--- ÉTAPE 4 : VÉRIFICATION LLM (Ollama / {OLLAMA_MODEL}) ---")
    print(f"  Paires à vérifier : {len(pairs_to_verify)} total | {already_cached} en cache | {len(new_pairs)} nouveaux appels")

    if new_pairs:
        def _verify(p):
            result = verify_match_ollama(p['kw'], p['label'], p['source'], p['etape'])
            return p['cache_key'], result

        with ThreadPoolExecutor(max_workers=OLLAMA_WORKERS) as executor:
            futures = {executor.submit(_verify, p): p for p in new_pairs}
            for future in tqdm(as_completed(futures), total=len(new_pairs), desc="LLM verify"):
                cache_key, result = future.result()
                llm_cache[cache_key] = result

        # Sauvegarde du cache après chaque batch
        with open(LLM_CACHE, 'wb') as f:
            pickle.dump(llm_cache, f)
        print(f"  Cache LLM sauvegardé ({len(llm_cache)} entrées)")

    # Construire les colonnes de résultat LLM dans df_results
    verdict_col    = []
    confidence_col = []
    reason_col     = []

    pair_lookup = {p['kw']: p for p in pairs_to_verify}

    for _, row in df_results.iterrows():
        kw    = row['Mot-clé original']
        etape = row['Etape_alignement']
        if etape == 'NON TROUVÉ' or kw not in pair_lookup:
            verdict_col.append('N/A')
            confidence_col.append('N/A')
            reason_col.append('')
        else:
            p = pair_lookup[kw]
            res = llm_cache.get(p['cache_key'], {'verdict': 'ERROR', 'confidence': 'N/A', 'reason': 'Non trouvé en cache'})
            verdict_col.append(res['verdict'])
            confidence_col.append(res['confidence'])
            reason_col.append(res['reason'])

    df_results = df_results.copy()
    df_results['LLM_Verdict']     = verdict_col
    df_results['LLM_Confidence']  = confidence_col
    df_results['LLM_Reason']      = reason_col

    return df_results


def export_llm_report(df_results):
    """
    Génère un rapport Excel détaillé de la vérification LLM,
    avec un onglet résumé et un onglet détaillé mot-clé par mot-clé.
    """
    # --- Onglet détaillé ---
    report_rows = []
    for _, row in df_results.iterrows():
        etape   = row['Etape_alignement']
        verdict = row.get('LLM_Verdict', 'N/A')

        # Récupérer label, URI, score/distance selon l'étape
        label, uri, source, score_info = '', '', '', ''
        if etape == 'Exact':
            label  = row.get(f'Label_{THRESHOLDS[0]}', '')
            uri    = row.get(f'ID_{THRESHOLDS[0]}', '')
            source = row.get(f'Source_{THRESHOLDS[0]}', '')
            score_info = '100.0 (exact)'
        elif etape == 'Fuzzy':
            for t in THRESHOLDS:
                if row.get(f'Source_{t}', 'NON TROUVÉ') != 'NON TROUVÉ':
                    label  = row.get(f'Label_{t}', '')
                    uri    = row.get(f'ID_{t}', '')
                    source = row.get(f'Source_{t}', '')
                    score_info = f"{row.get(f'Score_{t}', '')} (seuil {t})"
                    break
        elif etape == 'Sémantique':
            for thr in DIST_THRESHOLDS:
                col = str(thr).replace('.', '')
                if row.get(f'Source_sem_{col}', 'NON TROUVÉ') != 'NON TROUVÉ':
                    label  = row.get(f'Label_sem_{col}', '')
                    uri    = row.get(f'ID_sem_{col}', '')
                    source = row.get(f'Source_sem_{col}', '')
                    score_info = f"dist={row.get(f'Distance_sem_{col}', '')} (seuil {thr})"
                    break

        report_rows.append({
            'Mot-clé original':  row['Mot-clé original'],
            'Étape alignement':  etape,
            'Label trouvé':      label,
            'URI':               uri,
            'Vocabulaire':       source,
            'Score / Distance':  score_info,
            'LLM Verdict':       verdict,
            'LLM Confiance':     row.get('LLM_Confidence', 'N/A'),
            'LLM Explication':   row.get('LLM_Reason', ''),
        })

    df_detail = pd.DataFrame(report_rows)

    # --- Onglet résumé ---
    total        = len(df_results)
    non_trouves  = (df_results['Etape_alignement'] == 'NON TROUVÉ').sum()
    verifies     = (df_results['LLM_Verdict'] != 'N/A').sum()
    yes_high     = ((df_results['LLM_Verdict'] == 'YES') & (df_results['LLM_Confidence'] == 'HIGH')).sum()
    yes_medium   = ((df_results['LLM_Verdict'] == 'YES') & (df_results['LLM_Confidence'] == 'MEDIUM')).sum()
    yes_low      = ((df_results['LLM_Verdict'] == 'YES') & (df_results['LLM_Confidence'] == 'LOW')).sum()
    no_total     = (df_results['LLM_Verdict'] == 'NO').sum()
    errors       = (df_results['LLM_Verdict'] == 'ERROR').sum()

    summary_rows = [
        {'Métrique': 'Total mots-clés',                    'Valeur': total},
        {'Métrique': 'Non trouvés (aucune étape)',          'Valeur': int(non_trouves)},
        {'Métrique': 'Paires soumises au LLM',             'Valeur': int(verifies)},
        {'Métrique': 'Validés YES (confiance HIGH)',        'Valeur': int(yes_high)},
        {'Métrique': 'Validés YES (confiance MEDIUM)',      'Valeur': int(yes_medium)},
        {'Métrique': 'Validés YES (confiance LOW)',         'Valeur': int(yes_low)},
        {'Métrique': 'Rejetés NO',                         'Valeur': int(no_total)},
        {'Métrique': 'Erreurs LLM',                        'Valeur': int(errors)},
        {'Métrique': 'Taux validation (YES / vérifiés)',   'Valeur': f"{(yes_high+yes_medium+yes_low)/verifies*100:.1f}%" if verifies else "N/A"},
    ]

    # Résumé par étape
    etape_llm = []
    for etape in ['Exact', 'Fuzzy', 'Sémantique']:
        sub = df_results[df_results['Etape_alignement'] == etape]
        yes = (sub['LLM_Verdict'] == 'YES').sum()
        no  = (sub['LLM_Verdict'] == 'NO').sum()
        etape_llm.append({
            'Étape': etape,
            'Total': len(sub),
            'LLM YES': int(yes),
            'LLM NO':  int(no),
            'Taux validation': f"{yes/len(sub)*100:.1f}%" if len(sub) else "N/A"
        })

    with pd.ExcelWriter(FILE_LLM_REPORT, engine='openpyxl') as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name='Résumé', index=False)
        pd.DataFrame(etape_llm).to_excel(writer, sheet_name='Résumé par étape', index=False)
        df_detail.to_excel(writer, sheet_name='Détail mot-clé par mot-clé', index=False)

        # Mise en forme conditionnelle de l'onglet détail
        ws = writer.sheets['Détail mot-clé par mot-clé']
        from openpyxl.styles import PatternFill, Font
        green  = PatternFill("solid", fgColor="C6EFCE")
        red    = PatternFill("solid", fgColor="FFC7CE")
        orange = PatternFill("solid", fgColor="FFEB9C")
        bold   = Font(bold=True)

        # En-têtes en gras
        for cell in ws[1]:
            cell.font = bold

        # Colorisation par verdict
        verdict_col_idx = df_detail.columns.get_loc('LLM Verdict') + 1
        for row_idx, row_data in df_detail.iterrows():
            v = row_data['LLM Verdict']
            fill = green if v == 'YES' else (red if v == 'NO' else orange)
            ws.cell(row=row_idx + 2, column=verdict_col_idx).fill = fill

        # Largeur colonnes auto
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    print(f"  Rapport LLM exporté : {FILE_LLM_REPORT}")


# =============================================================================
# ÉTAPE 5 : LLM SUR MOT-CLÉS NON TROUVÉS (arbitrage parmi top-N candidats FAISS)
# =============================================================================

def verify_notfound_ollama(keyword, candidates):
    """
    Soumet un mot-clé NON TROUVÉ à Mistral avec ses top-N candidats FAISS.
    Retourne le meilleur label trouvé, ou AUCUN si aucun candidat ne convient.
    """
    candidates_txt = "\n".join([
        f"  {i+1}. \"{c['label']}\" ({c['source']}, distance={c['dist']})"
        for i, c in enumerate(candidates)
    ])

    prompt = f"""Tu es un expert en terminologie scientifique.
Le mot-clé suivant n'a pas pu être aligné automatiquement sur un vocabulaire contrôlé.

Mot-clé : "{keyword}"

Voici les {len(candidates)} candidats les plus proches trouvés par similarité vectorielle :
{candidates_txt}

Parmi ces candidats, y en a-t-il un qui correspond sémantiquement au mot-clé ?
Réponds UNIQUEMENT avec ces 3 lignes, sans aucun texte supplémentaire :
BEST_MATCH: le numéro du meilleur candidat (1 à {len(candidates)}), ou AUCUN si aucun ne convient
CONFIDENCE: HIGH ou MEDIUM ou LOW
REASON: une phrase expliquant ta décision"""

    try:
        response = requests.post(
            OLLAMA_URL,
            json={"model": OLLAMA_MODEL, "prompt": prompt, "stream": False},
            timeout=OLLAMA_TIMEOUT
        )
        if response.status_code != 200:
            return {'best_match': 'ERROR', 'confidence': 'N/A', 'reason': f'HTTP {response.status_code}',
                    'label': '', 'uri': '', 'source': ''}

        text = response.json().get("response", "").strip()

        best_match_raw = "AUCUN"
        confidence     = "UNKNOWN"
        reason         = ""

        for line in text.splitlines():
            line = line.strip()
            if line.upper().startswith("BEST_MATCH:"):
                best_match_raw = line.split(":", 1)[1].strip()
            elif line.upper().startswith("CONFIDENCE:"):
                val = line.split(":", 1)[1].strip().upper()
                if val == "HIGH":      confidence = "HIGH"
                elif val == "MEDIUM":  confidence = "MEDIUM"
                elif val == "LOW":     confidence = "LOW"
                elif "MEDIUM" in val:  confidence = "MEDIUM"
                elif "HIGH" in val:    confidence = "HIGH"
                elif "LOW" in val:     confidence = "LOW"
            elif line.upper().startswith("REASON:"):
                reason = line.split(":", 1)[1].strip()

        # Résoudre le numéro de candidat en label/uri/source
        if "AUCUN" in best_match_raw.upper():
            return {'best_match': 'AUCUN', 'confidence': confidence, 'reason': reason,
                    'label': '', 'uri': '', 'source': ''}

        try:
            idx = int(re.search(r'\d+', best_match_raw).group()) - 1
            if 0 <= idx < len(candidates):
                c = candidates[idx]
                return {'best_match': str(idx + 1), 'confidence': confidence, 'reason': reason,
                        'label': c['label'], 'uri': c['uri'], 'source': c['source']}
        except Exception:
            pass

        return {'best_match': 'PARSE_ERROR', 'confidence': confidence, 'reason': reason,
                'label': '', 'uri': '', 'source': ''}

    except requests.exceptions.Timeout:
        return {'best_match': 'ERROR', 'confidence': 'N/A', 'reason': 'Timeout Ollama',
                'label': '', 'uri': '', 'source': ''}
    except Exception as e:
        return {'best_match': 'ERROR', 'confidence': 'N/A', 'reason': str(e),
                'label': '', 'uri': '', 'source': ''}


def run_llm_step5(df_results, semantic_results, llm_cache_step5, master_index=None, index_faiss=None, labels_list=None, model=None):
    """
    Étape 5 : pour chaque mot-clé NON TROUVÉ aux étapes 1-3,
    soumet les top-N candidats FAISS à Mistral pour arbitrage.
    Si les candidats ne sont pas dans semantic_results (mots-clés jamais passés
    par l'Étape 3), ils sont calculés à la volée via FAISS.
    Enrichit le DataFrame avec les colonnes LLM5_* et retourne le rapport.
    """
    kws_not_found = [
        row['Mot-clé original']
        for _, row in df_results.iterrows()
        if row['Etape_alignement'] == 'NON TROUVÉ'
    ]

    # Calculer les candidats FAISS à la volée pour les mots sans candidats
    kws_need_faiss = [
        kw for kw in kws_not_found
        if not (semantic_results.get(kw) and semantic_results[kw].get('candidates'))
    ]

    if kws_need_faiss and index_faiss is not None and labels_list is not None and model is not None:
        print(f"  Calcul des candidats FAISS pour {len(kws_need_faiss)} mots NON TROUVÉS...")
        opt_batch = get_optimal_batch_size()
        kws_cleaned = [clean_term(k) for k in kws_need_faiss]
        embeddings  = model.encode(
            [BGE_QUERY_PREFIX + k for k in kws_cleaned],
            show_progress_bar=True, batch_size=opt_batch,
        )
        # Rapatrier sur CPU — gère tenseur torch, liste ou ndarray
        embeddings = to_numpy_cpu(embeddings)
        distances, indices = index_faiss.search(embeddings, k=TOP_N_CANDIDATES)

        for i, kw in enumerate(kws_need_faiss):
            candidates = []
            for rank in range(TOP_N_CANDIDATES):
                cidx = indices[i][rank]
                if cidx < len(labels_list):
                    candidates.append({
                        'label':  master_index[labels_list[cidx]][0],
                        'uri':    master_index[labels_list[cidx]][1],
                        'source': master_index[labels_list[cidx]][2],
                        'dist':   round(float(distances[i][rank]), 4),
                    })
            if kw not in semantic_results:
                semantic_results[kw] = {}
            semantic_results[kw]['candidates'] = candidates

        del embeddings; gc.collect()

    # Identifier les NON TROUVÉ qui ont des candidats FAISS
    pairs = []
    for _, row in df_results.iterrows():
        kw = row['Mot-clé original']
        if row['Etape_alignement'] != 'NON TROUVÉ':
            continue
        sem = semantic_results.get(kw)
        if sem and sem.get('candidates'):
            cache_key = f"step5||{kw}"
            pairs.append({'kw': kw, 'candidates': sem['candidates'], 'cache_key': cache_key})

    new_pairs    = [p for p in pairs if p['cache_key'] not in llm_cache_step5]
    cached_count = len(pairs) - len(new_pairs)

    print(f"\n--- ÉTAPE 5 : LLM SUR NON TROUVÉS ({OLLAMA_MODEL}) ---")
    print(f"  Mots-clés NON TROUVÉS avec candidats FAISS : {len(pairs)}")
    print(f"  En cache : {cached_count} | Nouveaux appels : {len(new_pairs)}")

    if new_pairs:
        def _verify5(p):
            result = verify_notfound_ollama(p['kw'], p['candidates'])
            return p['cache_key'], result

        with ThreadPoolExecutor(max_workers=OLLAMA_WORKERS) as executor:
            futures = {executor.submit(_verify5, p): p for p in new_pairs}
            for future in tqdm(as_completed(futures), total=len(new_pairs), desc="LLM step5"):
                cache_key, result = future.result()
                llm_cache_step5[cache_key] = result

        with open(LLM_CACHE_STEP5, 'wb') as f:
            pickle.dump(llm_cache_step5, f)
        print(f"  Cache Étape 5 sauvegardé ({len(llm_cache_step5)} entrées)")

    # Construire les colonnes LLM5 dans le DataFrame
    llm5_match      = []
    llm5_label      = []
    llm5_uri        = []
    llm5_source     = []
    llm5_confidence = []
    llm5_reason     = []

    pair_lookup = {p['kw']: p for p in pairs}

    for _, row in df_results.iterrows():
        kw = row['Mot-clé original']
        if row['Etape_alignement'] != 'NON TROUVÉ' or kw not in pair_lookup:
            llm5_match.append('N/A')
            llm5_label.append('')
            llm5_uri.append('')
            llm5_source.append('')
            llm5_confidence.append('N/A')
            llm5_reason.append('')
        else:
            p   = pair_lookup[kw]
            res = llm_cache_step5.get(p['cache_key'],
                  {'best_match': 'ERROR', 'confidence': 'N/A', 'reason': 'Non trouvé en cache',
                   'label': '', 'uri': '', 'source': ''})
            llm5_match.append(res['best_match'])
            llm5_label.append(res.get('label', ''))
            llm5_uri.append(res.get('uri', ''))
            llm5_source.append(res.get('source', ''))
            llm5_confidence.append(res['confidence'])
            llm5_reason.append(res['reason'])

    df_results = df_results.copy()
    df_results['LLM5_Match']      = llm5_match
    df_results['LLM5_Label']      = llm5_label
    df_results['LLM5_URI']        = llm5_uri
    df_results['LLM5_Source']     = llm5_source
    df_results['LLM5_Confidence'] = llm5_confidence
    df_results['LLM5_Reason']     = llm5_reason

    # Export rapport Étape 5
    export_llm_step5_report(df_results)

    return df_results


def export_llm_step5_report(df_results):
    """Génère le rapport Excel détaillé de l'Étape 5."""
    sub = df_results[df_results['Etape_alignement'] == 'NON TROUVÉ'].copy()

    report_rows = []
    for _, row in sub.iterrows():
        report_rows.append({
            'Mot-clé original':  row['Mot-clé original'],
            'LLM5 Match':        row.get('LLM5_Match', 'N/A'),
            'Label proposé':     row.get('LLM5_Label', ''),
            'URI':               row.get('LLM5_URI', ''),
            'Vocabulaire':       row.get('LLM5_Source', ''),
            'Confiance':         row.get('LLM5_Confidence', 'N/A'),
            'Explication':       row.get('LLM5_Reason', ''),
        })

    df_detail = pd.DataFrame(report_rows)

    # Garantir que les colonnes existent même si df_detail est vide
    for col in ['Mot-clé original', 'LLM5 Match', 'Label proposé', 'URI', 'Vocabulaire', 'Confiance', 'Explication']:
        if col not in df_detail.columns:
            df_detail[col] = []

    total   = len(sub)
    resolus = 0
    aucun   = 0
    errors  = 0
    if total > 0 and 'LLM5_Match' in sub.columns:
        resolus = (sub['LLM5_Match'].notna() & ~sub['LLM5_Match'].isin(['AUCUN', 'ERROR', 'N/A', 'PARSE_ERROR'])).sum()
        aucun   = (sub['LLM5_Match'] == 'AUCUN').sum()
        errors  = sub['LLM5_Match'].isin(['ERROR', 'PARSE_ERROR']).sum()

    summary_rows = [
        {'Métrique': 'Mots-clés NON TROUVÉS soumis',        'Valeur': int(total)},
        {'Métrique': 'Résolus par LLM (candidat retenu)',    'Valeur': int(resolus)},
        {'Métrique': 'Aucun candidat retenu (AUCUN)',        'Valeur': int(aucun)},
        {'Métrique': 'Erreurs LLM',                          'Valeur': int(errors)},
        {'Métrique': 'Taux de résolution',
         'Valeur': f"{resolus/total*100:.1f}%" if total else "N/A"},
    ]

    with pd.ExcelWriter(FILE_LLM_STEP5, engine='openpyxl') as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name='Résumé', index=False)
        df_detail.to_excel(writer, sheet_name='Détail NON TROUVÉS', index=False)

        ws = writer.sheets['Détail NON TROUVÉS']
        from openpyxl.styles import PatternFill, Font
        green  = PatternFill("solid", fgColor="C6EFCE")
        orange = PatternFill("solid", fgColor="FFEB9C")
        red    = PatternFill("solid", fgColor="FFC7CE")
        bold   = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold

        match_col_idx = df_detail.columns.get_loc('LLM5 Match') + 1
        for row_idx, row_data in df_detail.iterrows():
            v = row_data['LLM5 Match']
            if v not in ('AUCUN', 'ERROR', 'N/A', 'PARSE_ERROR'):
                fill = green
            elif v == 'AUCUN':
                fill = orange
            else:
                fill = red
            ws.cell(row=row_idx + 2, column=match_col_idx).fill = fill

        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    print(f"  Rapport Étape 5 exporté : {FILE_LLM_STEP5}")


# =============================================================================
# PIPELINE PRINCIPAL
# =============================================================================

def run():
    if not os.path.exists(VOCAB_DIR):
        os.makedirs(VOCAB_DIR)

    print_ram_usage("Démarrage")

    # =========================================================================
    # 1. GESTION INDEX LOCAL
    # =========================================================================
    master_index = {}
    rebuild_local = False

    if os.path.exists(INDEX_CACHE):
        if input("Mettre à jour l'index local ? (o/n) : ").lower() == 'o':
            rebuild_local = True
            # Invalider les caches dérivés
            for cache in [TFIDF_CACHE, FAISS_INDEX_PATH]:
                if os.path.exists(cache):
                    os.remove(cache)
                    print(f"  Cache invalidé : {os.path.basename(cache)}")
        else:
            with open(INDEX_CACHE, 'rb') as f:
                master_index = pickle.load(f)

    if rebuild_local or not master_index:
        print("\n--- INDEXATION LOCALE ---")
        for name, (fname, ftype) in tqdm(FILES.items()):
            if ftype == 'api': continue
            path = os.path.join(VOCAB_DIR, fname)
            if not os.path.exists(path): continue
            if ftype in ['rdf', 'xml']: master_index.update(parse_xml_robust(path, name))
            elif ftype == 'nq':         master_index.update(parse_mesh_nq(path))
            elif ftype == 'nt':         master_index.update(parse_agrovoc_nt(path))
            elif ftype == 'owl':        master_index.update(parse_obo_owl(path, name))
            elif ftype == 'sdf':        master_index.update(parse_chebi_sdf(path))
        if master_index:
            with open(INDEX_CACHE, 'wb') as f:
                pickle.dump(master_index, f)

    # =========================================================================
    # 2. CHARGEMENT MOTS-CLÉS
    # =========================================================================
    df_kw = pd.read_csv(FILE_KEYWORDS)
    kws   = [clean_keyword(k) for k in df_kw['keyword'].dropna().unique().tolist()]
    kws   = [k for k in kws if k]          # supprime les éventuelles chaînes vides
    kws   = list(dict.fromkeys(kws))        # dédoublonne après nettoyage (ex: "07. X" et "07.01 X" → même résultat)

    # =========================================================================
    # 3. GESTION CACHE PUBCHEM
    # =========================================================================
    pubchem_map = {}
    if os.path.exists(PUBCHEM_CACHE):
        if input("Réutiliser les résultats PubChem déjà téléchargés ? (o/n) : ").lower() == 'o':
            with open(PUBCHEM_CACHE, 'rb') as f:
                pubchem_map = pickle.load(f)

    kws_to_query = [k for k in kws if k not in pubchem_map]
    if kws_to_query:
        print(f"\n--- APPELS API PUBCHEM ({len(kws_to_query)} nouveaux) ---")
        with ThreadPoolExecutor(max_workers=3) as executor:
            future_to_kw = {executor.submit(query_pubchem, k): k for k in kws_to_query}
            for future in tqdm(as_completed(future_to_kw), total=len(kws_to_query), desc="PubChem"):
                res = future.result()
                pubchem_map[future_to_kw[future]] = res  # None si pas de résultat — évite de réinterroger l'API
        with open(PUBCHEM_CACHE, 'wb') as f:
            pickle.dump(pubchem_map, f)

    # =========================================================================
    # ÉTAPE 1 : EXACT MATCH
    # =========================================================================
    print(f"\n--- ÉTAPE 1 : EXACT MATCH ---")
    exact_results = {}     # kw -> {'label', 'uri', 'source'}
    kws_for_fuzzy = []

    for kw in kws:
        kw_s   = str(kw).strip()
        kw_l   = kw_s.lower()
        pc_uri = pubchem_map.get(kw_s)

        if kw_l in master_index:
            lbl, uid, src, _ = master_index[kw_l]
            exact_results[kw_s] = {'label': lbl, 'uri': uid, 'source': src}
        elif pc_uri:
            exact_results[kw_s] = {'label': kw_s, 'uri': pc_uri, 'source': 'PUBCHEM'}
        else:
            kws_for_fuzzy.append(kw_s)

    print(f"  Trouvés en exact : {len(exact_results)} | Restants pour étape 2 : {len(kws_for_fuzzy)}")

    # =========================================================================
    # ÉTAPE 2 : FUZZY (TF-IDF + RapidFuzz WRatio) — multi-processus
    # =========================================================================
    fuzzy_results = {}   # kw -> dict colonnes par seuil

    if kws_for_fuzzy:
        print(f"\n--- ÉTAPE 2 : FUZZY MATCH ({MAX_WORKERS} workers) ---")

        refs = list(master_index.keys())
        vectorizer = None

        if os.path.exists(TFIDF_CACHE):
            print("  Chargement du vectoriseur TF-IDF depuis le cache...")
            with open(TFIDF_CACHE, 'rb') as f:
                vectorizer = pickle.load(f)
        else:
            vectorizer = build_tfidf_vectorizer(refs)
            with open(TFIDF_CACHE, 'wb') as f:
                pickle.dump(vectorizer, f)
            print("  Vectoriseur TF-IDF sauvegardé.")

        kws_fuzzy_arr = np.array(kws_for_fuzzy)

        import platform
        use_multiprocess = platform.system() != 'Darwin' and MAX_WORKERS > 1

        if use_multiprocess:
            # Linux/Windows : multiprocessus via fork (initializer + globaux)
            with ProcessPoolExecutor(
                max_workers=MAX_WORKERS,
                initializer=_init_worker,
                initargs=(master_index, refs, vectorizer)
            ) as ex:
                futures = [
                    ex.submit(
                        match_worker_fuzzy,
                        kws_fuzzy_arr[i:i+500].tolist(),
                        THRESHOLDS,
                        pubchem_map
                    )
                    for i in range(0, len(kws_fuzzy_arr), 500)
                ]
                for f in tqdm(as_completed(futures), total=len(futures)):
                    fuzzy_results.update(f.result())
        else:
            # macOS (spawn) ou 1 worker : ThreadPoolExecutor — RapidFuzz libère le GIL
            print(f"  Mode threading (macOS/spawn-safe, {MAX_WORKERS} threads)")
            # Précalcul unique de la matrice sparse TF-IDF (500k+ termes × n-grammes).
            # Sans ça, vectorizer.transform(refs) est appelé dans chaque worker à
            # chaque chunk, soit 9× sur ce corpus — le vrai goulot de l'étape 2.
            print("  Précalcul de la matrice TF-IDF (opération unique)...")
            shared_ref_matrix = vectorizer.transform(refs)
            print(f"  Matrice TF-IDF prête : {shared_ref_matrix.shape[0]} termes × {shared_ref_matrix.shape[1]} n-grammes")
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
                futures = [
                    ex.submit(
                        match_worker_fuzzy,
                        kws_fuzzy_arr[i:i+500].tolist(),
                        THRESHOLDS,
                        pubchem_map,
                        master_index,       # passés directement — pas de globaux
                        refs,
                        vectorizer,
                        shared_ref_matrix   # matrice précalculée — partagée entre tous les chunks
                    )
                    for i in range(0, len(kws_fuzzy_arr), 500)
                ]
                for f in tqdm(as_completed(futures), total=len(futures)):
                    fuzzy_results.update(f.result())

        # Mots non trouvés même en fuzzy (score < seuil minimal pour tous les seuils)
        kws_for_semantic = [
            kw for kw in kws_for_fuzzy
            if fuzzy_results.get(kw, {}).get(f'Source_{FUZZY_THRESHOLD_FOR_SEMANTIC}', 'NON TROUVÉ') == 'NON TROUVÉ'
        ]
        found_fuzzy = len(kws_for_fuzzy) - len(kws_for_semantic)
        print(f"  Trouvés en fuzzy (seuil {min(THRESHOLDS)}) : {found_fuzzy} | Restants pour étape 3 : {len(kws_for_semantic)}")
        scores_98 = [
            fuzzy_results.get(kw, {}).get('Score_98', 0)
            for kw in kws_for_fuzzy
        ]
        print(f"  Score fuzzy_98 min={min(scores_98):.1f} max={max(scores_98):.1f} "
            f"moyenne={sum(scores_98)/len(scores_98):.1f}")
        print(f"  Mots avec Score_98 >= 98 : {sum(1 for s in scores_98 if s >= 98)}")
        print(f"  Mots avec Score_98 < 98  : {sum(1 for s in scores_98 if s < 98)}")
    else:
        kws_for_semantic = []

    # =========================================================================
    # ÉTAPE 3 : SÉMANTIQUE (BGE + FAISS)
    # Toujours initialiser model, index_faiss et labels_list — ils seront
    # réutilisés à l'Étape 5 même si aucun mot-clé ne passe par l'Étape 3.
    # =========================================================================
    semantic_results = {}   # kw -> {'label', 'uri', 'source', 'dist', 'candidates'}

    print(f"\n--- INITIALISATION BGE+FAISS (Étapes 3 & 5) ---")
    print_ram_usage("Avant chargement modèle")

    device = "mps" if torch.backends.mps.is_available() else "cpu"
    print(f" Device utilisé : {device}")
    model     = SentenceTransformer(MODEL_NAME, device=device)
    opt_batch = get_optimal_batch_size()
    print(f"  Modèle chargé. Batch size optimisé : {opt_batch}")

    labels_list = list(master_index.keys())

    if not os.path.exists(FAISS_INDEX_PATH):
        print("  Génération de l'index FAISS avec Normalisation L2...")
        # Sur M1/M2 16 Go, MPS sature rapidement la mémoire unifiée lors de
        # l'encodage en masse avec BGE-M3 (buffers intermédiaires 3-5× la taille
        # du batch). On force le device CPU pour l'encodage de l'index afin
        # d'éviter la memory pressure et le swap SSD qui font passer l'ETA de
        # ~5h à 50h+. Le device MPS est conservé pour les requêtes (opt_batch).
        # On force l'utilisation du GPU Apple (MPS)
        index_device = device 
        model_index = SentenceTransformer(MODEL_NAME, device=index_device)

        # On réduit drastiquement le batch pour ne pas saturer les 16 Go de RAM unifiée
        index_batch  = 32 
        print(f"  ⚡ Encodage index sur {index_device} avec batch_size={index_batch}")
        index_faiss  = None
        total_chunks = (len(labels_list) + index_batch - 1) // index_batch

        for chunk_i in tqdm(range(total_chunks), desc="  Encodage index FAISS"):
            chunk = labels_list[chunk_i * index_batch:(chunk_i + 1) * index_batch]
            chunk_emb = model_index.encode(
                chunk,
                show_progress_bar=False,
                batch_size=index_batch,
                convert_to_numpy=True,
            )
            # Rapatrier sur CPU — gère tenseur torch, liste ou ndarray
            chunk_emb = to_numpy_cpu(chunk_emb)
            # CRUCIAL : Normalisation L2 chunk par chunk
            faiss.normalize_L2(chunk_emb)
            if index_faiss is None:
                index_faiss = faiss.IndexFlatL2(chunk_emb.shape[1])
            index_faiss.add(chunk_emb)
            del chunk_emb   # libération immédiate via refcount — gc.collect() inutile ici

        del model_index   # libère la seconde instance du modèle après l'indexation

        faiss.write_index(index_faiss, FAISS_INDEX_PATH)
        print(f"  Index FAISS sauvegardé : {index_faiss.ntotal} vecteurs")
    else:
        index_faiss = faiss.read_index(FAISS_INDEX_PATH)

    if kws_for_semantic:
        print(f"\n--- ÉTAPE 3 : RECHERCHE SÉMANTIQUE BGE-M3+FAISS ({len(kws_for_semantic)} mots-clés) ---")

        # Nettoyage à la volée des préfixes numériques pour ne pas fausser le sens
        kws_sem_cleaned  = [re.sub(r'^\d+[\.\d+]*\s*', '', str(k)).strip() for k in kws_for_semantic]

        print("  Encodage des mots-clés restants...")
        kws_sem_embeddings = model.encode(
            [BGE_QUERY_PREFIX + k for k in kws_sem_cleaned],
            show_progress_bar=True, batch_size=opt_batch,
        )
        # Rapatrier sur CPU — gère tenseur torch, liste ou ndarray
        kws_sem_embeddings = to_numpy_cpu(kws_sem_embeddings)
        # CRUCIAL : Normalisation des vecteurs des requêtes
        faiss.normalize_L2(kws_sem_embeddings)

        distances, indices = index_faiss.search(kws_sem_embeddings, k=TOP_N_CANDIDATES)

        for i, kw in enumerate(kws_for_semantic):
            idx = indices[i][0]
            # Top-1 pour l'assemblage classique (Étape 3)
            semantic_results[kw] = {
                'label':  master_index[labels_list[idx]][0],
                'uri':    master_index[labels_list[idx]][1],
                'source': master_index[labels_list[idx]][2],
                'dist':   distances[i][0],
            }
            # Top-N candidats pour l'Étape 5 (LLM sur NON TROUVÉ)
            candidates = []
            for rank in range(TOP_N_CANDIDATES):
                cidx = indices[i][rank]
                if cidx < len(labels_list):
                    candidates.append({
                        'label':  master_index[labels_list[cidx]][0],
                        'uri':    master_index[labels_list[cidx]][1],
                        'source': master_index[labels_list[cidx]][2],
                        'dist':   round(float(distances[i][rank]), 4),
                    })
            semantic_results[kw]['candidates'] = candidates

        del kws_sem_embeddings; gc.collect()
        print_ram_usage("Après recherche sémantique")
    else:
        print("  Étape 3 : aucun mot-clé à traiter (tous résolus en exact/fuzzy).")

    # =========================================================================
    # ASSEMBLAGE DES RÉSULTATS
    # =========================================================================
    print("\n--- ASSEMBLAGE DES RÉSULTATS ---")
    final_data = []

    for kw in kws:
        kw_s = str(kw).strip()
        row  = {'Mot-clé original': kw_s}

        if kw_s in exact_results:
            # --- Étape 1 : exact match ---
            m = exact_results[kw_s]
            row['Etape_alignement'] = 'Exact'
            # Colonnes Étape 2 (fuzzy) : score 100 pour tous les seuils
            for t in THRESHOLDS:
                row[f'Label_{t}']  = m['label']
                row[f'ID_{t}']     = m['uri']
                row[f'Source_{t}'] = m['source']
                row[f'Score_{t}']  = 100.0
            # Colonnes Étape 3 (sémantique) : distance 0.0 pour tous les seuils
            for thr in DIST_THRESHOLDS:
                col = str(thr).replace('.', '')
                row[f'Label_sem_{col}']    = m['label']
                row[f'ID_sem_{col}']       = m['uri']
                row[f'Source_sem_{col}']   = m['source']
                row[f'Distance_sem_{col}'] = 0.0
        elif kw_s in fuzzy_results and fuzzy_results[kw_s].get(f'Score_{min(THRESHOLDS)}', 0) > 0:
            # --- Étape 2 : fuzzy ---
            row['Etape_alignement'] = 'Fuzzy'
            fz = fuzzy_results[kw_s]
            for t in THRESHOLDS:
                row[f'Label_{t}']  = fz[f'Label_{t}']
                row[f'ID_{t}']     = fz[f'ID_{t}']
                row[f'Source_{t}'] = fz[f'Source_{t}']
                row[f'Score_{t}']  = fz[f'Score_{t}']
            # Colonnes sémantiques : NON TROUVÉ (non traité à cette étape)
            for thr in DIST_THRESHOLDS:
                col = str(thr).replace('.', '')
                row[f'Label_sem_{col}']    = "NON TROUVÉ"
                row[f'ID_sem_{col}']       = ""
                row[f'Source_sem_{col}']   = "NON TROUVÉ"
                row[f'Distance_sem_{col}'] = 9.9

        elif kw_s in semantic_results:
            # --- Étape 3 : sémantique (L2 normalisé — distance inférieure = meilleur) ---
            sm = semantic_results[kw_s]
            row['Etape_alignement'] = 'Sémantique'
            # Colonnes fuzzy : NON TROUVÉ (non traité à cette étape)
            for t in THRESHOLDS:
                row[f'Label_{t}']  = "NON TROUVÉ"
                row[f'ID_{t}']     = ""
                row[f'Source_{t}'] = "NON TROUVÉ"
                row[f'Score_{t}']  = 0.0
            # Traiter les seuils sémantiques (distances L2 normalisées)
            is_semantic_success = False
            for thr in DIST_THRESHOLDS:
                col = str(thr).replace('.', '')
                # Avec L2 normé, on cherche une distance INFÉRIEURE OU ÉGALE au seuil
                if sm['dist'] <= thr:
                    row[f'Label_sem_{col}']    = sm['label']
                    row[f'ID_sem_{col}']       = sm['uri']
                    row[f'Source_sem_{col}']   = sm['source']
                    row[f'Distance_sem_{col}'] = round(sm['dist'], 4)
                    is_semantic_success = True
                else:
                    row[f'Label_sem_{col}']    = "NON TROUVÉ"
                    row[f'ID_sem_{col}']       = ""
                    row[f'Source_sem_{col}']   = "NON TROUVÉ"
                    row[f'Distance_sem_{col}'] = 9.9
            # Si le mot-clé ne passe aucun seuil (même le plus large), reclasser en NON TROUVÉ
            if not is_semantic_success:
                row['Etape_alignement'] = 'NON TROUVÉ'

        else:
            # Non trouvé à aucune étape
            row['Etape_alignement'] = 'NON TROUVÉ'
            for t in THRESHOLDS:
                row[f'Label_{t}']  = "NON TROUVÉ"
                row[f'ID_{t}']     = ""
                row[f'Source_{t}'] = "NON TROUVÉ"
                row[f'Score_{t}']  = 0.0
            for thr in DIST_THRESHOLDS:
                col = str(thr).replace('.', '')
                row[f'Label_sem_{col}']    = "NON TROUVÉ"
                row[f'ID_sem_{col}']       = ""
                row[f'Source_sem_{col}']   = "NON TROUVÉ"
                row[f'Distance_sem_{col}'] = 9.9

        final_data.append(row)

    df_res = pd.DataFrame(final_data)
    df_res.to_csv(FILE_OUTPUT_CSV, index=False, sep=';', encoding='utf-8-sig')
    print(f"  Export CSV : {FILE_OUTPUT_CSV}")

    # =========================================================================
    # ÉTAPE 4 : VÉRIFICATION LLM (Ollama / Mistral)
    # 
    # Installation :
    # Télécharger Ollama sur ollama.com
    # Lancer ollama pull mistral:7b dans ton terminal
    # Ollama tourne en local sur http://localhost:11434
    #
    # Bash
    # bashollama serve          # si pas déjà actif
    # ollama pull mistral:7b
    # =========================================================================
    llm_cache = {}
    if os.path.exists(LLM_CACHE):
        with open(LLM_CACHE, 'rb') as f:
            llm_cache = pickle.load(f)

    if check_ollama_available():
        df_res = run_llm_verification(df_res, llm_cache)
        export_llm_report(df_res)
        # Mise à jour du CSV avec les colonnes LLM
        df_res.to_csv(FILE_OUTPUT_CSV, index=False, sep=';', encoding='utf-8-sig')
        print(f"  CSV mis à jour avec colonnes LLM : {FILE_OUTPUT_CSV}")

        # =====================================================================
        # ÉTAPE 5 : LLM SUR MOT-CLÉS NON TROUVÉS
        # =====================================================================
        llm_cache_step5 = {}
        if os.path.exists(LLM_CACHE_STEP5):
            with open(LLM_CACHE_STEP5, 'rb') as f:
                llm_cache_step5 = pickle.load(f)

        df_res = run_llm_step5(df_res, semantic_results, llm_cache_step5,
                               master_index=master_index, index_faiss=index_faiss,
                               labels_list=labels_list, model=model)
        df_res.to_csv(FILE_OUTPUT_CSV, index=False, sep=';', encoding='utf-8-sig')
        print(f"  CSV mis à jour avec colonnes LLM5 : {FILE_OUTPUT_CSV}")
    else:
        print("  ⚠️  Étape 4 ignorée (Ollama non disponible). Relancez après : ollama serve && ollama pull mistral:7b")

    # =========================================================================
    # STATS — matrice vocabulaires × seuils (fuzzy + sémantique)
    # =========================================================================
    print("\n--- MATRICE DE STATISTIQUES ---")
    sources = ["NON TROUVÉ", "PUBCHEM"] + sorted([s for s in FILES.keys() if s != 'PUBCHEM'])

    counts_fuzzy = {t: df_res[f'Source_{t}'].value_counts() for t in THRESHOLDS}
    # Correction : on compte uniquement les lignes où la source n'est pas 'NON TROUVÉ',
    # ce qui respecte la condition de distance déjà appliquée lors de l'assemblage.
    counts_sem = {
        str(thr).replace('.', ''): df_res[f'Source_sem_{str(thr).replace(".", "")}'].value_counts()
        for thr in DIST_THRESHOLDS
    }

    matrix = []
    for src in sources:
        row = {'Vocabulaire': src}
        for t in THRESHOLDS:
            row[f'Fuzzy_Seuil_{t}'] = counts_fuzzy[t].get(src, 0)
        for thr in DIST_THRESHOLDS:
            col = str(thr).replace('.', '')
            row[f'Sem_Seuil_{thr}'] = counts_sem[col].get(src, 0)
        matrix.append(row)

    # Résumé par étape
    etape_counts = df_res['Etape_alignement'].value_counts().to_dict()
    etape_summary = [
        {'Etape': k, 'Nb_mots_clés': v}
        for k, v in etape_counts.items()
    ]

    with pd.ExcelWriter(FILE_STATS_XLSX, engine='openpyxl') as writer:
        pd.DataFrame(matrix).to_excel(writer, sheet_name='Matrice Vocabulaires', index=False)
        pd.DataFrame(etape_summary).to_excel(writer, sheet_name='Résumé Étapes', index=False)

    print(f"✅ Terminé : {FILE_STATS_XLSX}")
    print_ram_usage("Final")

# =============================================================================
# POINT D'ENTRÉE
# =============================================================================

if __name__ == "__main__":
    multiprocessing.freeze_support()
    run()